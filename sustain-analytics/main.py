import requests
import openpyxl
from bs4 import BeautifulSoup
import asyncio
from pyppeteer import launch

# excel=openpyxl.Workbook()
# # print(excel.sheetnames)
# sheet=excel.active
# sheet.title='Top Rated companies'
# # print(excel.sheetnames)
# sheet.append(['Company','Type','Rating','Review'])

async def main():
    browser = await launch()
    page = await browser.newPage()
    await page.goto('https://www.sustainalytics.com/esg-ratings')
    await page.waitFor(5000)

    ## Get HTML
    html = await page.content()
    await browser.close()
    return html

html_response = asyncio.get_event_loop().run_until_complete(main())

  
# defining the api-endpoint 
API_ENDPOINT = "https://www.sustainalytics.com/sustapi/companyratings/getcompanyratings"

  
# data to be sent to api
#for loop so that u can update the API data
for i in range(1, 6):  # Example: Loop from 1 to 5
    data = {
        "page": i,
        "pageSize": 10,
        "resourcePackage": "Sustainalytics"
    }
  
# sending post request and saving response as response object
r = requests.post(url = API_ENDPOINT, data = data)
  
# extracting response text 
data = r.text
print(data)

# Load HTML Response Into BeautifulSoup
soup = BeautifulSoup(html_response, "html.parser")
title = soup.find('section',id='company_ratings').find_all('div',class_='company-row')
# print('title', title)
# print(html_response)
# print(len(title))

for comp in title:
    name=comp.find('div',class_='w-50').a.text
    type=comp.find('div',class_='w-50').small.text
    compScore=comp.find('div',class_='col-2').text
    feedback=comp.find('div',class_='col-lg-6').text
    # print(name)
    # print(type)
    # print(compScore)
    # print(feedback)

#     sheet.append([name,type,compScore,feedback])
# excel.save('top ESG Comapany.xlsx')